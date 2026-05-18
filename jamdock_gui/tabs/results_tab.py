"""Tab 4 — Results & Analysis. Replaces ``jamrank``.

Features
--------
* Live updates: subscribes to ``DockingPool.job_finished`` from Tab 3 and
  appends rows as ligands finish docking.
* Bulk reload: rebuilds the table from ``docking_results/*.log`` on demand.
* Reactive filters: max affinity, min SimScore, min modes, MW range,
  ZINC-only, top-N. All filtering happens in the proxy — instant.
* Sortable columns: click any header to sort numerically.
* Exports: CSV, XLSX (if openpyxl), top-N poses zip, lab notebook .md,
  ZINC links txt.
"""
from __future__ import annotations

import csv
import zipfile
from datetime import datetime
from pathlib import Path

from PySide6.QtCore import Qt, Slot
from PySide6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QSplitter,
    QTableView,
    QVBoxLayout,
    QWidget,
)

from jamdock_gui.core.browser import open_url
from jamdock_gui.core.pymol_launcher import find_pymol, open_receptor_with_pose
from jamdock_gui.core.results import ResultRow, ResultsAggregator
from jamdock_gui.tabs.base_tab import BaseTab
from jamdock_gui.widgets.results_model import (
    COL_AFFINITY,
    COL_LIGAND,
    ResultsFilterProxy,
    ResultsTableModel,
)


class ResultsTab(BaseTab):
    def __init__(self, settings, parent: QWidget | None = None) -> None:
        super().__init__(settings, parent)
        self._aggregator = ResultsAggregator(self)
        self._model = ResultsTableModel(self)
        self._proxy = ResultsFilterProxy(self)
        self._proxy.setSourceModel(self._model)
        self._pool = None  # type: ignore[assignment]
        self._build()
        self._wire()
        self.refresh_from_workdir()

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------
    def _build(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)

        # -- header row: status + reload ---------------------------------
        head = QHBoxLayout()
        head.addWidget(QLabel("<b>Run:</b>"))
        self.cb_run = QComboBox()
        self.cb_run.addItem("(no docking_results/ folder detected)")
        self.cb_run.setEnabled(False)
        head.addWidget(self.cb_run, 1)

        self.lbl_live = QLabel("● Idle")
        self.lbl_live.setStyleSheet(
            "QLabel { color: #888; font-weight: bold; padding: 4px 10px; }"
        )
        head.addWidget(self.lbl_live)

        self.btn_reload = QPushButton("⟲  Reload")
        self.btn_reload.clicked.connect(self._on_reload)
        head.addWidget(self.btn_reload)
        root.addLayout(head)

        # -- filters ----------------------------------------------------
        filt_box = QGroupBox("Filters")
        f = QFormLayout(filt_box)

        self.sp_aff = QDoubleSpinBox()
        self.sp_aff.setRange(-30.0, 0.0); self.sp_aff.setSingleStep(0.5)
        self.sp_aff.setValue(-7.0); self.sp_aff.setSpecialValueText("(none)")
        self.sp_aff.setMinimum(-30.01)  # special value sentinel
        f.addRow("Max affinity (kcal/mol):", self.sp_aff)

        self.sp_sim = QSpinBox()
        self.sp_sim.setRange(-1, 100); self.sp_sim.setSuffix(" %")
        self.sp_sim.setValue(60); self.sp_sim.setSpecialValueText("(none)")
        f.addRow("Min SimScore:", self.sp_sim)

        self.sp_modes = QSpinBox()
        self.sp_modes.setRange(0, 50); self.sp_modes.setValue(5)
        self.sp_modes.setSpecialValueText("(none)")
        f.addRow("Min total modes:", self.sp_modes)

        mw_row = QHBoxLayout()
        self.sp_mw_min = QSpinBox(); self.sp_mw_min.setRange(0, 9999)
        self.sp_mw_min.setValue(0); self.sp_mw_min.setSpecialValueText("(none)")
        self.sp_mw_max = QSpinBox(); self.sp_mw_max.setRange(0, 9999)
        self.sp_mw_max.setValue(0); self.sp_mw_max.setSpecialValueText("(none)")
        mw_row.addWidget(self.sp_mw_min); mw_row.addWidget(QLabel("–"))
        mw_row.addWidget(self.sp_mw_max); mw_row.addWidget(QLabel("Da"))
        mw_row.addStretch(1)
        mw_w = QWidget(); mw_w.setLayout(mw_row)
        f.addRow("MW range:", mw_w)

        self.cb_only_zinc = QCheckBox("Only entries with a ZINC ID")
        f.addRow("", self.cb_only_zinc)

        self.cb_only_ro5 = QCheckBox(
            "Only Ro5-compliant (drug-like — MW≤500, LogP≤5, HBD≤5, HBA≤10)"
        )
        f.addRow("", self.cb_only_ro5)

        self.sp_top = QSpinBox(); self.sp_top.setRange(0, 1_000_000)
        self.sp_top.setValue(50); self.sp_top.setSpecialValueText("(all)")
        f.addRow("Show top:", self.sp_top)

        btns = QHBoxLayout()
        self.btn_apply = QPushButton("Apply filters")
        self.btn_reset = QPushButton("⟲  Reset")
        self.btn_apply.clicked.connect(self._apply_filters)
        self.btn_reset.clicked.connect(self._reset_filters)
        btns.addWidget(self.btn_apply); btns.addWidget(self.btn_reset)
        btns.addStretch(1)
        wbtns = QWidget(); wbtns.setLayout(btns)
        f.addRow("", wbtns)
        root.addWidget(filt_box)

        # -- table ------------------------------------------------------
        splitter = QSplitter(Qt.Horizontal)
        self.table = QTableView()
        self.table.setModel(self._proxy)
        self.table.setSortingEnabled(True)
        self.table.sortByColumn(COL_AFFINITY, Qt.AscendingOrder)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(COL_LIGAND, QHeaderView.Stretch)
        splitter.addWidget(self.table)

        # -- sidebar (status + future 3D viewer placeholder) -----------
        side = QWidget()
        sv = QVBoxLayout(side)
        sv.setContentsMargins(8, 0, 0, 0)
        self.lbl_count = QLabel("0 rows · filtered: 0")
        self.lbl_count.setStyleSheet("QLabel { color: #555; padding: 4px; }")
        sv.addWidget(self.lbl_count)

        sv.addWidget(QLabel("<b>Selected pose</b>"))
        self.lbl_selected = QLabel("(click a row)")
        self.lbl_selected.setWordWrap(True)
        self.lbl_selected.setTextFormat(Qt.RichText)
        self.lbl_selected.setTextInteractionFlags(
            Qt.TextBrowserInteraction | Qt.LinksAccessibleByMouse
        )
        # Don't let Qt try to handle the URL itself — its detector fails on
        # WSL2 with "Unable to detect a web browser". We route via our own
        # core.browser helper which knows about wslview / explorer.exe etc.
        self.lbl_selected.setOpenExternalLinks(False)
        self.lbl_selected.linkActivated.connect(self._on_link_activated)
        self.lbl_selected.setStyleSheet(
            "QLabel { background: #f8f8f8; border: 1px solid #e0e0e0;"
            " padding: 8px; border-radius: 3px; }"
        )
        sv.addWidget(self.lbl_selected)

        # External PyMOL viewer — useful when the embedded NGL viewer is
        # unavailable (e.g. WSL2). Launches PyMOL with receptor + pose loaded.
        self.btn_open_pymol = QPushButton("Open pose in PyMOL")
        self.btn_open_pymol.setEnabled(False)
        self.btn_open_pymol.clicked.connect(self._on_open_pymol)
        sv.addWidget(self.btn_open_pymol)

        sv.addStretch(1)
        side.setMaximumWidth(320)
        splitter.addWidget(side)
        splitter.setStretchFactor(0, 4)
        splitter.setStretchFactor(1, 1)
        root.addWidget(splitter, 1)

        # -- exports ----------------------------------------------------
        exp_row = QHBoxLayout()
        exp_row.addWidget(QLabel("<b>Export filtered:</b>"))
        self.btn_csv = QPushButton("CSV")
        self.btn_xlsx = QPushButton("XLSX")
        self.btn_zip = QPushButton("Top poses (zip)")
        self.btn_md = QPushButton("Lab notebook (.md)")
        self.btn_zinc = QPushButton("ZINC links (.txt)")
        for b in (self.btn_csv, self.btn_xlsx, self.btn_zip, self.btn_md, self.btn_zinc):
            exp_row.addWidget(b)
            b.setEnabled(False)
        exp_row.addStretch(1)
        root.addLayout(exp_row)

        self.btn_csv.clicked.connect(self._export_csv)
        self.btn_xlsx.clicked.connect(self._export_xlsx)
        self.btn_zip.clicked.connect(self._export_poses_zip)
        self.btn_md.clicked.connect(self._export_notebook)
        self.btn_zinc.clicked.connect(self._export_zinc_links)

    def _wire(self) -> None:
        self._aggregator.row_added.connect(self._on_row_added)
        self._aggregator.row_updated.connect(self._on_row_updated)
        self._aggregator.bulk_loaded.connect(self._on_bulk_loaded)
        self.table.selectionModel().selectionChanged.connect(self._on_row_selected)
        self._proxy.layoutChanged.connect(self._refresh_count_label)
        self._proxy.rowsInserted.connect(self._refresh_count_label)
        self._proxy.rowsRemoved.connect(self._refresh_count_label)
        self._proxy.modelReset.connect(self._refresh_count_label)

    # ------------------------------------------------------------------
    # Workdir wiring
    # ------------------------------------------------------------------
    def refresh_from_workdir(self) -> None:
        wd = self.workdir
        self._aggregator.set_workdir(wd)
        self._update_run_label()
        if wd:
            self._aggregator.scan_workdir()
        else:
            self._model.replace_all([])

    def _update_run_label(self) -> None:
        wd = self.workdir
        self.cb_run.clear()
        if not wd:
            self.cb_run.addItem("(no working directory)")
            self.cb_run.setEnabled(False)
            return
        results_dir = wd / "docking_results"
        if results_dir.is_dir():
            n = sum(1 for _ in results_dir.glob("*.log"))
            self.cb_run.addItem(f"{results_dir.name} — {n} logs")
            self.cb_run.setEnabled(True)
        else:
            self.cb_run.addItem("(no docking_results/ folder yet)")
            self.cb_run.setEnabled(False)

    # ------------------------------------------------------------------
    # Live pool wiring (called by MainWindow)
    # ------------------------------------------------------------------
    @Slot(object)
    def attach_pool(self, pool) -> None:
        self._pool = pool
        # New jobs: each job_finished(idx) → ingest the corresponding log.
        pool.job_finished.connect(self._on_pool_job_finished)
        pool.all_finished.connect(self._on_pool_all_finished)
        self._set_live_indicator(True)

    @Slot()
    def detach_pool(self) -> None:
        if self._pool is not None:
            try:
                self._pool.job_finished.disconnect(self._on_pool_job_finished)
            except (RuntimeError, TypeError):
                pass
            try:
                self._pool.all_finished.disconnect(self._on_pool_all_finished)
            except (RuntimeError, TypeError):
                pass
        self._pool = None
        self._set_live_indicator(False)

    @Slot(int)
    def _on_pool_job_finished(self, idx: int) -> None:
        if not self._pool:
            return
        try:
            job = self._pool.jobs[idx]
        except IndexError:
            return
        self._aggregator.ingest_log(job.log_path)

    @Slot()
    def _on_pool_all_finished(self) -> None:
        # Pool ran to completion; do a final scan to pick up any rows that
        # may have raced (unlikely but cheap insurance).
        self._aggregator.scan_workdir()
        self._set_live_indicator(False)

    def _set_live_indicator(self, live: bool) -> None:
        if live:
            self.lbl_live.setText("● Live")
            self.lbl_live.setStyleSheet(
                "QLabel { color: #1e8449; font-weight: bold;"
                " padding: 4px 10px; }"
            )
        else:
            self.lbl_live.setText("● Idle")
            self.lbl_live.setStyleSheet(
                "QLabel { color: #888; font-weight: bold; padding: 4px 10px; }"
            )

    # ------------------------------------------------------------------
    # Aggregator → model
    # ------------------------------------------------------------------
    @Slot(object)
    def _on_row_added(self, row: ResultRow) -> None:
        self._model.upsert(row)
        self._enable_exports()

    @Slot(object)
    def _on_row_updated(self, row: ResultRow) -> None:
        self._model.upsert(row)

    @Slot(int)
    def _on_bulk_loaded(self, n: int) -> None:
        self._model.replace_all(self._aggregator.rows)
        self._enable_exports()
        self._update_run_label()

    def _on_reload(self) -> None:
        self._aggregator.scan_workdir()

    # ------------------------------------------------------------------
    # Filters
    # ------------------------------------------------------------------
    def _apply_filters(self) -> None:
        # Map the spin "(none)" sentinels back to None.
        aff = self.sp_aff.value()
        self._proxy.set_max_affinity(None if aff <= -30.0 else aff)
        sim = self.sp_sim.value()
        self._proxy.set_min_sim_score(None if sim < 0 else sim)
        modes = self.sp_modes.value()
        self._proxy.set_min_modes(None if modes <= 0 else modes)

        mw_min = self.sp_mw_min.value() or None
        mw_max = self.sp_mw_max.value() or None
        self._proxy.set_mw_range(mw_min, mw_max)
        self._proxy.set_only_with_zinc(self.cb_only_zinc.isChecked())
        self._proxy.set_only_ro5(self.cb_only_ro5.isChecked())
        top = self.sp_top.value()
        self._proxy.set_top_n(None if top <= 0 else top)

    def _reset_filters(self) -> None:
        self._proxy.reset_filters()
        self.sp_aff.setValue(-30.01)   # → "(none)"
        self.sp_sim.setValue(-1)
        self.sp_modes.setValue(0)
        self.sp_mw_min.setValue(0)
        self.sp_mw_max.setValue(0)
        self.cb_only_zinc.setChecked(False)
        self.cb_only_ro5.setChecked(False)
        self.sp_top.setValue(0)

    # ------------------------------------------------------------------
    # Selection feedback
    # ------------------------------------------------------------------
    @Slot()
    def _on_row_selected(self) -> None:
        row = self._selected_row()
        if not row:
            self.lbl_selected.setText("(click a row)")
            self.btn_open_pymol.setEnabled(False)
            return
        # Enable PyMOL button if PyMOL is on PATH and we have a pose file.
        self.btn_open_pymol.setEnabled(
            row.pose_path.is_file() and find_pymol() is not None
        )
        link = (
            f"<a href='{row.zinc_link}'>{row.zinc_id}</a>"
            if row.zinc_link else (row.zinc_id or "—")
        )
        aff_s = f"{row.affinity:.2f}" if row.affinity is not None else "—"
        mw_s = f"{row.mw:.1f}" if row.mw is not None else "—"
        sim_s = f"{row.sim_score} %" if row.sim_score is not None else "—"
        self.lbl_selected.setText(
            f"<b>{row.ligand}</b><br>"
            f"Affinity: <b>{aff_s}</b> kcal/mol<br>"
            f"SimScore: {sim_s}<br>"
            f"Modes: {row.n_modes}<br>"
            f"MW: {mw_s} Da<br>"
            f"ZINC: {link}<br>"
            f"<small>pose: {row.pose_path.name}</small>"
        )

    def _selected_row(self) -> ResultRow | None:
        sel = self.table.selectionModel().selectedRows()
        if not sel:
            return None
        proxy_idx = sel[0]
        src_idx = self._proxy.mapToSource(proxy_idx)
        return self._model.row_at(src_idx.row())

    # ------------------------------------------------------------------
    # Bookkeeping
    # ------------------------------------------------------------------
    def _refresh_count_label(self) -> None:
        total = self._model.rowCount()
        shown = self._proxy.rowCount()
        self.lbl_count.setText(f"{total} rows · filtered: {shown}")
        self._enable_exports()

    def _enable_exports(self) -> None:
        n = self._proxy.rowCount()
        for b in (self.btn_csv, self.btn_xlsx, self.btn_zip,
                  self.btn_md, self.btn_zinc):
            b.setEnabled(n > 0)

    # ------------------------------------------------------------------
    # Exports — operate on the *filtered/sorted* visible rows
    # ------------------------------------------------------------------
    def _visible_rows(self) -> list[ResultRow]:
        """Return rows in the proxy's current order (filtered + sorted)."""
        n = self._proxy.rowCount()
        out: list[ResultRow] = []
        for r in range(n):
            src_idx = self._proxy.mapToSource(self._proxy.index(r, 0))
            row = self._model.row_at(src_idx.row())
            if row is not None:
                out.append(row)
        return out

    def _timestamp(self) -> str:
        return datetime.now().strftime("%Y_%m_%d_%H%M")

    def _ask_save(self, default_name: str, fmt_label: str) -> Path | None:
        start = str((self.workdir or Path.home()) / default_name)
        chosen, _ = QFileDialog.getSaveFileName(self, "Save export", start, fmt_label)
        return Path(chosen) if chosen else None

    def _export_csv(self) -> None:
        path = self._ask_save(f"top_hits_{self._timestamp()}.csv", "CSV (*.csv)")
        if not path:
            return
        rows = self._visible_rows()
        with path.open("w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["rank", "ligand", "affinity_kcal_mol", "sim_score_pct",
                        "n_modes", "mw_da", "logp", "hbd", "hba",
                        "ro5_violations", "ro5_pass",
                        "zinc_id", "zinc_link", "pose_path", "note"])
            for i, r in enumerate(rows, 1):
                w.writerow([i, r.ligand, r.affinity, r.sim_score, r.n_modes,
                            r.mw, r.logp, r.hbd, r.hba,
                            r.ro5_violations,
                            "" if r.ro5_pass is None else int(r.ro5_pass),
                            r.zinc_id or "", r.zinc_link or "",
                            str(r.pose_path), r.error or ""])
        QMessageBox.information(self, "CSV exported", f"Saved {len(rows)} rows to:\n{path}")

    def _export_xlsx(self) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            QMessageBox.warning(
                self, "openpyxl not installed",
                "Install <code>openpyxl</code> for XLSX export:<br>"
                "<code>pip install openpyxl</code>"
            )
            return
        path = self._ask_save(f"top_hits_{self._timestamp()}.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        rows = self._visible_rows()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Top hits"
        headers = ["Rank", "Ligand", "Affinity (kcal/mol)", "SimScore (%)",
                   "Modes", "MW (Da)", "LogP", "HBD", "HBA",
                   "Ro5 violations", "Drug-like",
                   "ZINC ID", "ZINC link", "Pose", "Note"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for i, r in enumerate(rows, 1):
            ws.append([i, r.ligand, r.affinity, r.sim_score, r.n_modes,
                       r.mw, r.logp, r.hbd, r.hba,
                       r.ro5_violations,
                       "" if r.ro5_pass is None else ("yes" if r.ro5_pass else "no"),
                       r.zinc_id, r.zinc_link, str(r.pose_path), r.error])
        for col_idx, h in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, len(h) + 2)
        wb.save(path)
        QMessageBox.information(self, "XLSX exported", f"Saved {len(rows)} rows to:\n{path}")

    def _export_poses_zip(self) -> None:
        path = self._ask_save(f"top_poses_{self._timestamp()}.zip", "Zip (*.zip)")
        if not path:
            return
        rows = self._visible_rows()
        n_added = 0
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
            for r in rows:
                if r.pose_path and r.pose_path.is_file():
                    zf.write(r.pose_path, arcname=r.pose_path.name)
                    n_added += 1
        QMessageBox.information(self, "Poses exported",
                                f"Wrote {n_added} pose files to:\n{path}")

    def _export_notebook(self) -> None:
        path = self._ask_save(f"lab_notebook_{self._timestamp()}.md",
                              "Markdown (*.md)")
        if not path:
            return
        rows = self._visible_rows()
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        lines = [
            "# Docking results — lab notebook",
            "",
            f"*Generated by jamdock-gui on {ts}*",
            "",
            f"Working directory: `{self.workdir}`",
            f"Total rows in run: {self._model.rowCount()}",
            f"Rows after filters / top-N: {len(rows)}",
            "",
            "## Top hits",
            "",
            "| Rank | Ligand | Affinity | SimScore | Modes "
            "| MW | LogP | HBD | HBA | Drug-like | ZINC ID |",
            "|---:|---|---:|---:|---:|---:|---:|---:|---:|:---:|---|",
        ]
        for i, r in enumerate(rows, 1):
            zinc = (
                f"[{r.zinc_id}]({r.zinc_link})"
                if (r.zinc_id and r.zinc_link) else (r.zinc_id or "—")
            )
            aff_s = f"{r.affinity:.2f}" if r.affinity is not None else "—"
            sim_s = str(r.sim_score) if r.sim_score is not None else "—"
            mw_s = f"{r.mw:.1f}" if r.mw is not None else "—"
            lp_s = f"{r.logp:.2f}" if r.logp is not None else "—"
            hbd_s = str(r.hbd) if r.hbd is not None else "—"
            hba_s = str(r.hba) if r.hba is not None else "—"
            ro5_s = "—"
            if r.ro5_pass is True:
                ro5_s = "★"
            elif r.ro5_pass is False:
                ro5_s = f"{r.ro5_violations} viol"
            lines.append(
                f"| {i} | {r.ligand} | {aff_s} | {sim_s} | {r.n_modes} "
                f"| {mw_s} | {lp_s} | {hbd_s} | {hba_s} | {ro5_s} | {zinc} |"
            )
        lines += [
            "",
            "## Notes",
            "",
            "_Add your interpretation here._",
            "",
            "---",
            "Citations: jamdock-suite (10.5281/zenodo.15577778), "
            "QuickVina 2 (10.1093/bioinformatics/btv082), "
            "AutoDock Vina (10.1002/jcc.21334).",
        ]
        path.write_text("\n".join(lines), encoding="utf-8")
        QMessageBox.information(self, "Notebook exported", f"Saved to:\n{path}")

    @Slot(str)
    def _on_link_activated(self, url: str) -> None:
        """Open *url* in the system browser, with WSL fallbacks.

        Bypasses Qt's broken-on-WSL ``QDesktopServices.openUrl`` by using
        the multi-backend helper in :mod:`core.browser`.

        Also copies the URL to the clipboard on every click. ZINC's web
        server occasionally returns a 500 on the first "cold" hit; having
        the URL already on the clipboard lets the user paste it into a
        fresh tab and retry without going back through the table.
        """
        # Stash the URL on the clipboard before opening, so a cold-500
        # from zinc.docking.org can be retried with a quick Ctrl-V in a
        # new tab. This does clobber the clipboard's prior contents, but
        # the alternative (only copying on failure) means the user has
        # to navigate back here to recover from a transient ZINC error.
        from PySide6.QtGui import QGuiApplication
        cb = QGuiApplication.clipboard()
        if cb is not None:
            cb.setText(url)

        if open_url(url):
            return
        # If every method failed, the URL is already on the clipboard.
        QMessageBox.information(
            self, "Browser not detected",
            "Could not detect a web browser to open the link. "
            "The URL has been copied to your clipboard:\n\n" + url,
        )

    def _on_open_pymol(self) -> None:
        # Launch PyMOL with the receptor + the selected pose loaded.
        row = self._selected_row()
        if row is None:
            return
        wd = self.workdir
        if not wd:
            return
        pymol = find_pymol()
        pymol = find_pymol()
        if not pymol:
            QMessageBox.warning(self, "PyMOL not found",
                "PyMOL is not on PATH. Install it and relaunch.")
            return
        receptors = sorted(wd.glob("*_for_docking*.pdbqt"))
        if not receptors:
            QMessageBox.warning(self, "Receptor not found",
                "No receptor PDBQT in working directory.")
            return
        ok, msg = open_receptor_with_pose(
            pymol=pymol, receptor=receptors[-1],
            pose=row.pose_path, workdir=wd,
        )
        if not ok:
            QMessageBox.warning(self, "PyMOL launch failed", msg)

    def _export_zinc_links(self) -> None:
        path = self._ask_save(f"zinc_links_{self._timestamp()}.txt", "Text (*.txt)")
        if not path:
            return
        rows = self._visible_rows()
        links = [r.zinc_link for r in rows if r.zinc_link]
        path.write_text("\n".join(links) + "\n", encoding="utf-8")
        QMessageBox.information(self, "Links exported",
            f"Wrote {len(links)} ZINC links to:\n{path}")
